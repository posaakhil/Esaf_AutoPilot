import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
import json

# ===== LOAD CONFIG FROM JSON =====
with open("esaf_config.json", 'r', encoding='utf-8') as f:
    config = json.load(f)

ASSIGNEES = config["assignees"]

# ===== CONFIG =====
AUTO_MODE = "--auto" in sys.argv

def load_data():
    """Load HCA_India, HCA_Domestic, and HCA_EXTRA_DATA (if exists) — SAFE"""
    if not os.path.exists("Today_Assignment.xlsx"):
        print("[ERROR] Today_Assignment.xlsx not found. Run step3_assign_split.py first.")
        return None, None, None

    try:
        wb = load_workbook("Today_Assignment.xlsx")
        sheet_names = wb.sheetnames

        # Load HCA_India (or empty)
        if "HCA_India" in sheet_names:
            india_df = pd.read_excel("Today_Assignment.xlsx", sheet_name="HCA_India")
        else:
            india_df = pd.DataFrame()

        # Load HCA_Domestic (or empty)
        if "HCA_Domestic" in sheet_names:
            domestic_df = pd.read_excel("Today_Assignment.xlsx", sheet_name="HCA_Domestic")
        else:
            domestic_df = pd.DataFrame()

        # Load HCA_EXTRA_DATA (or empty)
        if "HCA_EXTRA_DATA" in sheet_names:
            extra_df = pd.read_excel("Today_Assignment.xlsx", sheet_name="HCA_EXTRA_DATA")
        else:
            extra_df = pd.DataFrame()

        print(f"[INFO] Loaded  India={len(india_df)}, Domestic={len(domestic_df)}, Extra={len(extra_df)}")
        return india_df, domestic_df, extra_df

    except Exception as e:
        print(f"[ERROR] Failed to load: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

def create_summary_sheet(india_df, domestic_df, extra_df):
    """Create Summary sheet — works even if all are empty"""
    summary_data = []

    total_india = len(india_df)
    total_domestic = len(domestic_df)
    total_extra = len(extra_df)
    total_all = total_india + total_domestic + total_extra

    summary_data.append(["TOTAL REQUESTS", "India", "Domestic", "Extra", "All"])
    summary_data.append(["Count", total_india, total_domestic, total_extra, total_all])

    summary_data.append([])
    summary_data.append(["ASSIGNEE LOAD", "India", "Domestic", "Total"])
    for assignee in ASSIGNEES:
        india_count = len(india_df[india_df['Assignee'] == assignee]) if not india_df.empty else 0
        domestic_count = len(domestic_df[domestic_df['Assignee'] == assignee]) if not domestic_df.empty else 0
        total = india_count + domestic_count
        summary_data.append([assignee, india_count, domestic_count, total])

    summary_data.append([])
    summary_data.append(["APPLICATIONS", "Create", "Modify", "Delete", "Total"])

    all_apps = set()
    if not india_df.empty and 'Application' in india_df.columns:
        all_apps.update(india_df['Application'].dropna().unique())
    if not domestic_df.empty and 'Application' in domestic_df.columns:
        all_apps.update(domestic_df['Application'].dropna().unique())

    create_total = 0
    modify_total = 0
    delete_total = 0
    grand_total = 0

    for app in sorted(all_apps):
        create_count = 0
        modify_count = 0
        delete_count = 0

        if not india_df.empty and 'Request' in india_df.columns:
            app_mask = india_df['Application'] == app
            for req in india_df[app_mask]['Request'].dropna():
                if 'Create' in req:
                    create_count += 1
                elif 'Modify' in req:
                    modify_count += 1
                elif 'Delete' in req:
                    delete_count += 1

        if not domestic_df.empty and 'Request' in domestic_df.columns:
            app_mask = domestic_df['Application'] == app
            for req in domestic_df[app_mask]['Request'].dropna():
                if 'Create' in req:
                    create_count += 1
                elif 'Modify' in req:
                    modify_count += 1
                elif 'Delete' in req:
                    delete_count += 1

        total_app = create_count + modify_count + delete_count
        summary_data.append([app, create_count, modify_count, delete_count, total_app])
        create_total += create_count
        modify_total += modify_count
        delete_total += delete_count
        grand_total += total_app

    summary_data.append(["GrandFull Total", create_total, modify_total, delete_total, grand_total])
    return pd.DataFrame(summary_data[1:], columns=summary_data[0])

def create_pivot_tables(india_df, domestic_df):
    """Create Pivot Tables — handles empty DataFrames"""
    all_assignees = ASSIGNEES

    # India Pivot
    if not india_df.empty and 'Application' in india_df.columns and 'Assignee' in india_df.columns:
        india_pivot = india_df.pivot_table(
            index='Application',
            columns='Assignee',
            aggfunc='size',
            fill_value=0
        ).reindex(columns=all_assignees, fill_value=0)
        india_pivot['Grand Total'] = india_pivot.sum(axis=1)
        india_pivot = india_pivot.reset_index()
    else:
        india_pivot = pd.DataFrame(columns=['Application'] + all_assignees + ['Grand Total'])

    # Domestic Pivot
    if not domestic_df.empty and 'Application' in domestic_df.columns and 'Assignee' in domestic_df.columns:
        domestic_pivot = domestic_df.pivot_table(
            index='Application',
            columns='Assignee',
            aggfunc='size',
            fill_value=0
        ).reindex(columns=all_assignees, fill_value=0)
        domestic_pivot['Grand Total'] = domestic_pivot.sum(axis=1)
        domestic_pivot = domestic_pivot.reset_index()
    else:
        domestic_pivot = pd.DataFrame(columns=['Application'] + all_assignees + ['Grand Total'])

    return india_pivot, domestic_pivot

def apply_styling(worksheet, title=""):
    header_fill = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    if title:
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

    header_row = 2 if title else 1
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for row in range(header_row + 1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_fit_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        width = max(max_length + 2, 25) if column_letter == 'A' else min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = width

def save_to_sheets(summary_df, india_pivot, domestic_pivot):
    with pd.ExcelWriter("Today_Assignment.xlsx", engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        combined_pivot = pd.concat([india_pivot, domestic_pivot], ignore_index=True)
        combined_pivot.to_excel(writer, sheet_name="Pivot", index=False)

    wb = load_workbook("Today_Assignment.xlsx")

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        apply_styling(ws, "EXECUTIVE SUMMARY")
        auto_fit_columns(ws)
        last_row = ws.max_row
        for row in range(3, last_row + 1):
            ws.cell(row=row, column=2).font = Font(color="006400")  # Create
            ws.cell(row=row, column=3).font = Font(color="0000FF")  # Modify
            ws.cell(row=row, column=4).font = Font(color="FF0000")  # Delete

    if "Pivot" in wb.sheetnames:
        ws = wb["Pivot"]
        apply_styling(ws, "HCA INDIA & DOMESTIC PIVOT TABLES")
        auto_fit_columns(ws)

        last_row = ws.max_row
        last_col = ws.max_column
        if last_row >= 2:  # Only add total if there's data
            grand_total_row = last_row + 2
            ws.cell(row=grand_total_row, column=1, value="GrandFull Total")
            total_sum = 0
            for col in range(2, last_col):
                col_sum = sum(ws.cell(row=r, column=col).value or 0 for r in range(2, last_row + 1))
                ws.cell(row=grand_total_row, column=col, value=col_sum)
                total_sum += col_sum
            ws.cell(row=grand_total_row, column=last_col, value=total_sum)

            for col in range(1, last_col + 1):
                cell = ws.cell(row=grand_total_row, column=col)
                cell.font = Font(bold=True, color="2E5984")
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb.save("Today_Assignment.xlsx")
    print("[SUCCESS] Saved and styled Summary and Pivot sheets.")

def main():
    print("[INFO] STEP 4: EXECUTIVE SUMMARY & PIVOT TABLES")
    print("=" * 50)

    india_df, domestic_df, extra_df = load_data()
    if india_df is None:
        return

    summary_df = create_summary_sheet(india_df, domestic_df, extra_df)
    india_pivot, domestic_pivot = create_pivot_tables(india_df, domestic_df)
    save_to_sheets(summary_df, india_pivot, domestic_pivot)

    print("\n[SUCCESS] STEP 4 COMPLETED — EXECUTIVE DASHBOARD READY!")
    print("[INFO] Summary and Pivot sheets created successfully.")

if __name__ == "__main__":
    main()