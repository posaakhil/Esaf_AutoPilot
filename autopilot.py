import subprocess
import sys
import os
import json
import shutil
import keyboard
from colorama import Fore, Style, init
import pyautogui
import tempfile
import importlib.util

# Initialize colorama with full compatibility
init(autoreset=True, convert=True, strip=False)

CONFIG_FILE = "esaf_config.json"
DEFAULTS_FILE = "esaf_config_defaults.json"
BACKUP_FILE = "esaf_config_backup.json"

# ===== SCRIPT FILENAMES =====
SCRIPTS = {
    "step1": "esaf_automation.py",
    "step2": "merge_and_cleanup.py", 
    "step3": "Data_Analysis_Split.py",
    "step4": "summary_pivot.py",
    "step5": "Interactive_Dashboard.py",
    "complete": "complete_process.py"
}

def get_ascii_banner():
    """Return ASCII banner as plain string (safe for all terminals)"""
    return f"""
{Fore.CYAN}+==============================================================================+
{Fore.CYAN}|                                                                              |
{Fore.CYAN}|  {Fore.BLUE}███████╗███████╗ █████╗ ███████╗     █████╗ ██╗   ██╗████████╗ ██████╗ {Fore.CYAN}  |
{Fore.CYAN}|  {Fore.BLUE}██╔════╝██╔════╝██╔══██╗██╔════╝    ██╔══██╗██║   ██║╚══██╔══╝██╔═══██╗{Fore.CYAN}  |
{Fore.CYAN}|  {Fore.BLUE}█████╗  ███████╗███████║█████╗      ███████║██║   ██║   ██║   ██║   ██║{Fore.CYAN}  |
{Fore.CYAN}|  {Fore.BLUE}██╔══╝  ╚════██║██╔══██║██╔══╝      ██╔══██║██║   ██║   ██║   ██║   ██║{Fore.CYAN}  |
{Fore.CYAN}|  {Fore.BLUE}███████╗███████║██║  ██║██          ██║  ██║╚██████╔╝   ██║   ╚██████╔╝{Fore.CYAN}  |
{Fore.CYAN}|  {Fore.BLUE}╚══════╝╚══════╝╚═╝  ╚═╝╚═          ╚═╝  ╚═╝ ╚═════╝    ╚═╝    ╚═════╝ {Fore.CYAN}  |
{Fore.CYAN}|                                                                              |
{Fore.CYAN}|  {Fore.GREEN}+--------------------------------------------------------------------------+  {Fore.CYAN}|
{Fore.CYAN}|  {Fore.GREEN}|           AUTOMATED REQUEST MANAGEMENT SYSTEM - ESAF AutoPilot™         |  {Fore.CYAN}|
{Fore.CYAN}|  {Fore.GREEN}+--------------------------------------------------------------------------+  {Fore.CYAN}|
{Fore.CYAN}|                                                                              |
{Fore.CYAN}+==============================================================================+
"""

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def extract_script_to_temp(script_name):
    """Extract script from bundle to temporary file and return path"""
    try:
        # Get the bundled script
        bundled_path = get_resource_path(script_name)
        
        if not os.path.exists(bundled_path):
            print(f"{Fore.RED}[ERROR] Script not found in bundle: {script_name}")
            return None
        
        # Read the script content
        with open(bundled_path, 'r', encoding='utf-8') as f:
            script_content = f.read()
        
        # Create temporary file
        temp_dir = tempfile.mkdtemp()
        temp_script_path = os.path.join(temp_dir, script_name)
        
        # Write script to temporary file
        with open(temp_script_path, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        return temp_script_path
        
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Failed to extract script: {e}")
        return None

def run_script_directly(script_name, step_name):
    """Run script directly without creating new process"""
    try:
        print(f"\n{Fore.CYAN}[RUN] Starting {step_name}...")
        print(f"{Fore.YELLOW}[ABORT] Press ESC anytime to stop!")
        
        # Extract script to temp file
        script_path = extract_script_to_temp(script_name)
        if not script_path:
            return False
        
        # Read and execute the script
        with open(script_path, 'r', encoding='utf-8') as f:
            script_code = f.read()
        
        # Create a temporary module and execute it
        temp_module_name = f"temp_{script_name.replace('.', '_')}"
        spec = importlib.util.spec_from_loader(temp_module_name, loader=None)
        temp_module = importlib.util.module_from_spec(spec)
        
        # Set up the environment for the script
        script_globals = {
            '__name__': '__main__',
            '__file__': script_path,
            'sys': sys,
            'os': os,
            'json': json,
            'shutil': shutil,
            'keyboard': keyboard,
            'Fore': Fore,
            'Style': Style,
            'init': init,
            'pyautogui': pyautogui,
            'pd': None,  # pandas
            'load_workbook': None,  # openpyxl
            'px': None,  # plotly express
            'go': None,  # plotly graph_objects
            'pio': None,  # plotly io
            'subprocess': subprocess
        }
        
        # Try to import required modules for the script
        try:
            import pandas as pd
            script_globals['pd'] = pd
        except ImportError:
            pass
            
        try:
            from openpyxl import load_workbook
            script_globals['load_workbook'] = load_workbook
        except ImportError:
            pass
            
        try:
            import plotly.express as px
            import plotly.graph_objects as go
            import plotly.io as pio
            script_globals['px'] = px
            script_globals['go'] = go
            script_globals['pio'] = pio
        except ImportError:
            pass
        
        # Execute the script
        exec(script_code, script_globals)
        
        print(f"\n{Fore.GREEN}[SUCCESS] {step_name} completed!")
        return True
        
    except KeyboardInterrupt:
        print(f"\n{Fore.RED}[ABORT] Process interrupted by user.")
        return False
    except Exception as e:
        print(f"\n{Fore.RED}[ERROR] Failed to run {step_name}: {e}")
        return False

def check_abort():
    """Check if user pressed ESC - can be called anywhere"""
    if keyboard.is_pressed('esc'):
        print(f"\n{Fore.RED}[ABORT] EMERGENCY STOP: ESC key pressed!")
        raise KeyboardInterrupt("User pressed ESC")

def ensure_defaults():
    """Create defaults file if it doesn't exist"""
    if not os.path.exists(DEFAULTS_FILE):
        defaults = {
            "url": "https://esaf.hca.corpad.net/",
            "downloads_folder": "AUTO",
            "mouse_coords": {
                "requests_assigned": [168, 212],
                "advanced_filter": [728, 317],
                "app_field": [589, 516],
                "click_outside": [534, 512],
                "apply_button": [1357, 780],
                "export_xls": [1834, 317],
                "status_field": [588, 674],
                "apply_2_button": [1358, 781]
            },
            "timings": {
                "wait_after_page_load": 6,
                "wait_after_requests_click": 3,
                "wait_after_advanced_filter_open": 2,
                "wait_after_paste": 2,
                "wait_after_apply": 5,
                "wait_after_export": 3,
                "max_wait_for_download": 30,
                "check_interval": 1
            },
            "queues": ["CORP-ACCESS-INDIA"],
            "keywords": ["SMART", "SMART Inventory", "Smart admin", "abc", "HOST"],
            "assignees": ["Akhil", "Swathi", "Divya", "Amreen", "Riya", "Madhurima", "Vamsitha"],
            "rules": {
                "india_overflow_threshold": 70,
                "max_per_person_domestic": 15,
                "max_per_person_meditech": 5,
                "requests_per_app_domestic": 2,
                "india_keyword": "CORP-ACCESS-INDIA",
                "meditech_keyword": "MEDITECH_Expanse_CAP_Panhandle_Market"
            },
            "output_file": "Today_Assignment.xlsx",
            "keep_columns": [
                "User name",
                "sAMAccountName",
                "Request date",
                "Last updated time",
                "Requested by",
                "Request",
                "Status",
                "Application"
            ],
            "excel_options": {
                "auto_fit_column_width": True,
                "wrap_text": True
            },
            "no_requests_text": "There are no requests available",
            "base_assignment_folder": "assignment_"
        }
        with open(DEFAULTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(defaults, f, indent=4, ensure_ascii=False)
        print(f"{Fore.GREEN}[OK] Created default config: {DEFAULTS_FILE}")

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return None
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Failed to load config: {e}")
        return None

def resolve_downloads_folder(config):
    raw = config.get("downloads_folder", "AUTO")
    if raw == "AUTO" or not str(raw).strip():
        return os.path.join(os.path.expanduser("~"), "Downloads").replace("\\", "/")
    return str(raw).replace("\\", "/")

def save_config(config):
    if os.path.exists(CONFIG_FILE):
        shutil.copy2(CONFIG_FILE, BACKUP_FILE)
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)
    print(f"{Fore.GREEN}[SAVE] Config saved to {CONFIG_FILE}")

def reset_to_defaults():
    if os.path.exists(DEFAULTS_FILE):
        with open(DEFAULTS_FILE, 'r', encoding='utf-8') as f:
            defaults = json.load(f)
        save_config(defaults)
        print(f"{Fore.GREEN}[OK] Config reset to defaults!")
    else:
        print(f"{Fore.YELLOW}[WARN] Defaults file not found. Creating...")
        ensure_defaults()
        reset_to_defaults()

def validate_config():
    config = load_config()
    if not config:
        return False
    required_fields = ['url', 'downloads_folder', 'mouse_coords', 'keywords', 'queues', 'assignees', 'rules']
    for field in required_fields:
        if field not in config:
            print(f"{Fore.RED}[MISSING] Required field: {field}")
            return False
    mouse_coords = config.get('mouse_coords', {})
    required_coords = ['requests_assigned', 'advanced_filter', 'app_field', 'status_field',
                      'apply_button', 'apply_2_button', 'export_xls', 'click_outside']
    for coord in required_coords:
        if coord not in mouse_coords or not isinstance(mouse_coords[coord], list) or len(mouse_coords[coord]) != 2:
            print(f"{Fore.RED}[INVALID] Mouse coord: {coord}")
            return False
    return True

def edit_url(config):
    current = config.get("url", "https://esaf.hca.corpad.net/")
    print(f"\n{Fore.CYAN}[URL] Current: {Fore.WHITE}{current}")
    new_url = input(f"{Fore.CYAN}[EDIT] Enter new ESAF URL (or press Enter to keep): ").strip()
    final_url = new_url if new_url else current
    config["url"] = final_url
    save_config(config)
    print(f"{Fore.GREEN}[OK] URL updated to: {Fore.WHITE}{final_url}")

def capture_mouse_coordinates(config):
    print(f"\n{Fore.CYAN}[MOUSE] Coordinate Capture Tool")
    print(f"{Fore.YELLOW}[NOTE] Set browser zoom to 100% and keep window position fixed!")
    if "mouse_coords" not in config:
        config["mouse_coords"] = {}
    steps = [
        ("Requests Assigned to Me", "requests_assigned"),
        ("Advanced Filter Button", "advanced_filter"),
        ("Application Name Field", "app_field"),
        ("Apply Button", "apply_button"),
        ("Export XLS Button", "export_xls"),
        ("Click Outside Target", "click_outside"),
        ("Status Field", "status_field"),
        ("Apply_2 Button", "apply_2_button")
    ]
    for display_name, key in steps:
        input(f"\n{Fore.CYAN}[ACTION] Hover over '{display_name}', then press ENTER...")
        x, y = pyautogui.position()
        config["mouse_coords"][key] = [x, y]
        print(f"{Fore.GREEN}[CAPTURED] {key} -> ({x}, {y})")
    save_config(config)

def edit_list(config, key, name):
    current = config.get(key, [])
    print(f"\n{Fore.CYAN}[LIST] Current {name}: {Fore.WHITE}{current}")
    print(f"{Fore.YELLOW}1) Add item(s)")
    print(f"{Fore.YELLOW}2) Remove item(s)")
    print(f"{Fore.YELLOW}3) Clear all")
    print(f"{Fore.YELLOW}4) Back")
    choice = input(f"{Fore.CYAN}Choose option (1-4): ").strip()
    if choice == "1":
        try:
            count = int(input(f"{Fore.CYAN}How many {name.lower()} to add? ").strip())
            if count <= 0:
                print(f"{Fore.YELLOW}[SKIP] No items added.")
                return
            new_items = []
            for i in range(count):
                item = input(f"{Fore.CYAN}Enter {name[:-1].lower()} #{i+1}: ").strip()
                if item:
                    new_items.append(item)
            if new_items:
                current.extend(new_items)
                config[key] = current
                save_config(config)
                print(f"{Fore.GREEN}[OK] Added {len(new_items)} {name.lower()}.")
        except ValueError:
            print(f"{Fore.RED}[ERROR] Invalid number.")
    elif choice == "2" and current:
        print(f"\n{Fore.CYAN}[REMOVE] Select items to remove (comma-separated numbers):")
        for i, item in enumerate(current, 1):
            print(f"{Fore.YELLOW}{i}) {item}")
        indices_input = input(f"{Fore.CYAN}Enter numbers (e.g., 1,3,5) or 0 to cancel: ").strip()
        if indices_input == "0":
            return
        try:
            indices = [int(x.strip()) - 1 for x in indices_input.split(",") if x.strip().isdigit()]
            indices = sorted(set(indices), reverse=True)
            removed = []
            for idx in indices:
                if 0 <= idx < len(current):
                    removed.append(current.pop(idx))
            if removed:
                config[key] = current
                save_config(config)
                print(f"{Fore.GREEN}[OK] Removed: {removed}")
            else:
                print(f"{Fore.YELLOW}[WARN] No valid items removed.")
        except Exception as e:
            print(f"{Fore.RED}[ERROR] Invalid input.")
    elif choice == "3":
        config[key] = []
        save_config(config)
        print(f"{Fore.GREEN}[OK] Cleared all {name}")

def edit_rules(config):
    rules = config.get("rules", {})
    print(f"\n{Fore.CYAN}[RULES] Current Rules:")
    for key, value in rules.items():
        print(f"{Fore.YELLOW} - {key}: {value}")
    fields = {"1": "india_overflow_threshold", "2": "max_per_person_domestic", "3": "max_per_person_meditech", "4": "requests_per_app_domestic"}
    print(f"\n{Fore.YELLOW}1) Edit India Overflow Threshold")
    print(f"{Fore.YELLOW}2) Edit Max Per Person Domestic")
    print(f"{Fore.YELLOW}3) Edit Max Per Person Meditech")
    print(f"{Fore.YELLOW}4) Edit Requests Per App Domestic")
    print(f"{Fore.YELLOW}5) Back")
    choice = input(f"{Fore.CYAN}Choose option (1-5): ").strip()
    if choice in fields:
        try:
            value = int(input(f"{Fore.CYAN}Enter new value for {fields[choice]}: ").strip())
            rules[fields[choice]] = value
            config["rules"] = rules
            save_config(config)
            print(f"{Fore.GREEN}[OK] Rule updated!")
        except ValueError:
            print(f"{Fore.RED}[ERROR] Invalid number.")

def view_config(config):
    print(f"\n{Fore.CYAN}[CONFIG] CURRENT CONFIGURATION:")
    print(json.dumps(config, indent=2, ensure_ascii=False))

def show_menu():
    try:
        terminal_width = shutil.get_terminal_size().columns
    except:
        terminal_width = 80
    os.system('cls' if os.name == 'nt' else 'clear')
    print(get_ascii_banner())
    print()
    config_status = "[OK] CONFIGURED" if validate_config() else "[ERROR] NOT CONFIGURED"
    status_color = Fore.GREEN if validate_config() else Fore.RED
    status_line = f"[SYSTEM STATUS: {status_color}{config_status}{Fore.RESET}]"
    print(status_line.center(terminal_width))
    print()
    notices = [
        f"{Fore.CYAN}[TIP] First-time user? Configure (Option 1) before running full process!",
        f"{Fore.RED}[WARNING] Press ESC anytime during automation to ABORT immediately!",
        f"{Fore.GREEN}[INFO] HCA Healthcare Internal Use Only • v2.1"
    ]
    for notice in notices:
        print(notice.center(terminal_width))
    print()
    print(f"{Fore.MAGENTA} MAIN OPERATIONS MENU ".center(terminal_width, "="))
    print()
    menu_items = [
        f"{Fore.WHITE}1) {Fore.CYAN}Configuration Management",
        f"{Fore.WHITE}2) {Fore.GREEN}Run FULL End-to-End Process (complete_process.py)",
        f"{Fore.WHITE}3) {Fore.BLUE}Step 1: ESAF UI Automation (esaf_automation.py)",
        f"{Fore.WHITE}4) {Fore.BLUE}Step 2: Merge & Cleanup (merge_and_cleanup.py)",
        f"{Fore.WHITE}5) {Fore.BLUE}Step 3: Assign Requests to Team (Data_Analysis_Split.py)",
        f"{Fore.WHITE}6) {Fore.BLUE}Step 4: Summary + Pivot (summary_pivot.py)",
        f"{Fore.WHITE}7) {Fore.BLUE}Step 5: Interactive Dashboard (Interactive_Dashboard.py)",
        f"{Fore.WHITE}0) {Fore.RED}Exit ESAF AutoPilot"
    ]
    for item in menu_items:
        print("    " + item)
    print()
    footer = f"{Fore.CYAN}ESAF AutoPilot™ • Streamlining Access Request Management"
    print(footer.center(terminal_width))

def crud_menu():
    """Configuration Management Menu"""
    config = load_config() or {}
    while True:
        print(f"\n{Fore.CYAN}[CONFIG] CONFIGURATION MANAGEMENT")
        print(f"{Fore.CYAN}=" * 60)
        print(f"{Fore.YELLOW}0) {Fore.CYAN}Edit ESAF URL")
        print(f"{Fore.YELLOW}1) {Fore.CYAN}Capture Mouse Coordinates")
        print(f"{Fore.YELLOW}2) {Fore.CYAN}Edit Keywords")
        print(f"{Fore.YELLOW}3) {Fore.CYAN}Edit Assignees")
        print(f"{Fore.YELLOW}4) {Fore.CYAN}Set Downloads Folder")
        print(f"{Fore.YELLOW}5) {Fore.CYAN}Edit Queues")
        print(f"{Fore.YELLOW}6) {Fore.CYAN}Edit Rules (Thresholds)")
        print(f"{Fore.YELLOW}7) {Fore.CYAN}Reset to Defaults")
        print(f"{Fore.YELLOW}8) {Fore.CYAN}View Current Config")
        print(f"{Fore.YELLOW}9) {Fore.RED}Back to Main Menu")
        print(f"{Fore.CYAN}=" * 60)
        choice = input(f"{Fore.CYAN}Choose option (0-9): ").strip()
        if choice == "0":
            edit_url(config)
        elif choice == "1":
            capture_mouse_coordinates(config)
        elif choice == "2":
            edit_list(config, "keywords", "Keywords")
        elif choice == "3":
            edit_list(config, "assignees", "Assignees")
        elif choice == "4":
            folder = input(f"{Fore.CYAN}Enter Downloads folder path (or 'AUTO'): ").strip()
            config["downloads_folder"] = folder if folder else "AUTO"
            save_config(config)
        elif choice == "5":
            edit_list(config, "queues", "Queues")
        elif choice == "6":
            edit_rules(config)
        elif choice == "7":
            reset_to_defaults()
            config = load_config() or {}
        elif choice == "8":
            view_config(config)
        elif choice == "9":
            break
        else:
            print(f"{Fore.YELLOW}[WARN] Invalid choice.")

def main():
    ensure_defaults()
    while True:
        try:
            show_menu()
            choice = input(f"\n{Fore.CYAN}Enter your choice (0-7): ").strip()
            if choice == "1":
                crud_menu()
            elif choice == "2":
                run_script_directly(SCRIPTS["complete"], "Full End-to-End Process")
            elif choice == "3":
                run_script_directly(SCRIPTS["step1"], "Step 1: ESAF UI Automation")
            elif choice == "4":
                run_script_directly(SCRIPTS["step2"], "Step 2: Merge & Cleanup")
            elif choice == "5":
                run_script_directly(SCRIPTS["step3"], "Step 3: Assign Requests to Team")
            elif choice == "6":
                run_script_directly(SCRIPTS["step4"], "Step 4: Summary + Pivot")
            elif choice == "7":
                run_script_directly(SCRIPTS["step5"], "Step 5: Interactive Dashboard")
            elif choice == "0":
                print(f"\n{Fore.CYAN}Thank you for using ESAF AutoPilot™. Goodbye!")
                break
            else:
                print(f"{Fore.YELLOW}[WARN] Please enter 0-7.")
            
            if choice != "0":
                input(f"\n{Fore.CYAN}Press Enter to return to main menu...")
                
        except KeyboardInterrupt:
            print(f"\n{Fore.RED}[ABORT] Interrupted. Returning to menu...")
            input(f"\n{Fore.CYAN}Press Enter to continue...")
        except Exception as e:
            print(f"\n{Fore.RED}[ERROR] Unexpected error: {e}")
            input(f"\n{Fore.CYAN}Press Enter to continue...")

if __name__ == "__main__":
    # Ensure UTF-8 and color support
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding='utf-8')
    init(autoreset=True, convert=True, strip=False)
    main()
