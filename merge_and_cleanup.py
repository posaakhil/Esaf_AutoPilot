import os
import glob
import sys
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import subprocess

# ===== LOAD CONFIG FROM JSON =====
with open("esaf_config.json", 'r', encoding='utf-8') as f:
    config = json.load(f)

KEEP_COLUMNS = config["keep_columns"]
OUTPUT_FILE = config["output_file"]
AUTO_FIT_COLUMN_WIDTH = True
AUTO_MODE = "--auto" in sys.argv

def clean_value(value):
    """Convert to string and remove ALL non-ASCII characters"""
    if pd.isna(value):
        return ""
    if not isinstance(value, str):
        value = str(value)
    # Replace common Unicode
    replacements = {
        '\u2192': '->', '\u2190': '<-', '\u2191': '^', '\u2193': 'v',
        '\u2013': '-', '\u2014': '--', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\xa0': ' ',
        '\u200b': '', '\ufeff': '', '\u00a0': ' '
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    # Keep only printable ASCII (32–126)
    return ''.join(c for c in value if 32 <= ord(c) <= 126)

def get_all_assignment_folders():
    folders = [f for f in glob.glob("assignment_*") if os.path.isdir(f)]
    valid = []
    for f in folders:
        match = re.match(r'assignment_(\d+)$', f)
        if match:
            valid.append((int(match.group(1)), f))
    valid.sort()
    return [f for _, f in valid]

def select_folder_interactive(folders):
    if not folders:
        print("[ERROR] No assignment_X folders found.")
        return None
    print(f"\n[INFO] Found {len(folders)} assignment folders:")
    for i, folder in enumerate(folders, 1):
        count = len(glob.glob(os.path.join(folder, "*.xls")))
        print(f"  {i}) {folder} ({count} files)")
    while True:
        try:
            choice = input("\n[INPUT] Enter folder number (or 0 to exit): ").strip()
            if choice == "0":
                return None
            idx = int(choice) - 1
            if 0 <= idx < len(folders):
                return folders[idx]
            print("[ERROR] Invalid number.")
        except ValueError:
            print("[ERROR] Enter a number.")

def merge_excel_files(folder_name):
    files = glob.glob(os.path.join(folder_name, "*.xls"))
    if not files:
        print(f"[INFO] No .xls files in '{folder_name}'")
        return None

    print(f"[INFO] Found {len(files)} files. Merging...")
    dfs = []
    for file in files:
        try:
            df = pd.read_excel(file, dtype=str)
            # Clean EVERY cell in EVERY column
            for col in df.columns:
                df[col] = df[col].apply(clean_value)
            df['SourceFile'] = clean_value(os.path.basename(file))
            df['RunFolder'] = clean_value(folder_name)
            dfs.append(df)
            print(f"   -> Loaded: {os.path.basename(file)} ({len(df)} rows)")
        except Exception as e:
            print(f"   [ERROR] Failed to load {file}: {e}")
    if not dfs:
        return None
    merged = pd.concat(dfs, ignore_index=True)
    print(f"[SUCCESS] Merged {len(merged)} rows from '{folder_name}'.")
    return merged

def cleanup_columns(df):
    available = [col for col in KEEP_COLUMNS if col in df.columns]
    missing = [col for col in KEEP_COLUMNS if col not in df.columns]
    if missing:
        print(f"[WARNING] Missing columns: {missing}")
    cleaned = df[available].copy()
    print(f"[INFO] Final columns: {available}")
    return cleaned

def apply_excel_formatting(worksheet):
    worksheet.title = "Master_Data"
    # Header styling
    header_fill = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Sort by 'Last updated time'
    sort_col = None
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=col).value == "Last updated time":
            sort_col = col
            break

    if sort_col:
        print("[INFO] Sorting by 'Last updated time' (A to Z)...")
        rows = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
            rows.append(tuple(clean_value(cell) for cell in row))
        try:
            rows.sort(key=lambda x: x[sort_col - 1] or "")
        except Exception as e:
            print(f"[WARNING] Sort failed: {e}")
        # Clear and rewrite with cleaned values
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.value = None
        for r_idx, row_data in enumerate(rows, start=2):
            for c_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Auto-fit columns
    if AUTO_FIT_COLUMN_WIDTH:
        for col in worksheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            width = min(max_len + 2, 50)
            worksheet.column_dimensions[letter].width = width

def save_to_excel(df, append_mode=False):
    try:
        if append_mode and os.path.exists(OUTPUT_FILE):
            existing = pd.read_excel(OUTPUT_FILE, dtype=str)
            for col in existing.columns:
                existing[col] = existing[col].apply(clean_value)
            combined = pd.concat([existing, df], ignore_index=True)
            print(f"[INFO] Appended {len(df)} rows.")
        else:
            combined = df
            print("[INFO] Creating new file.")

        final_df = combined[KEEP_COLUMNS]
        # Final clean before save
        for col in final_df.columns:
            final_df[col] = final_df[col].apply(clean_value)
        final_df.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')

        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        apply_excel_formatting(ws)
        wb.save(OUTPUT_FILE)

        print(f"[SUCCESS] Saved: {OUTPUT_FILE}")
        print(f"[INFO] Shape: {final_df.shape[0]} rows, {final_df.shape[1]} cols")
        print("[INFO] Sheet: 'Master_Data' | Styled | Sorted")

    except Exception as e:
        # Avoid printing e if it contains Unicode
        print(f"[ERROR] Save failed: {str(e).encode('ascii', 'ignore').decode('ascii')}")

def main():
    print("[INFO] STEP 2: MERGE & CLEANUP EXCEL FILES — EXECUTIVE EDITION")
    print("=" * 50)

    folders = get_all_assignment_folders()
    if AUTO_MODE:
        if not folders:
            print("[ERROR] No folders in auto-mode.")
            return
        folder = folders[-1]
        print(f"[AUTO] Using: {folder}")
    else:
        if not folders:
            folder = input("[INPUT] Enter folder name: ").strip()
            if not os.path.exists(folder):
                print(f"[ERROR] Folder not found.")
                return
        else:
            folder = select_folder_interactive(folders)
            if not folder:
                return

    df = merge_excel_files(folder)
    if df is None:
        return

    cleaned_df = cleanup_columns(df)
    save_to_excel(cleaned_df, append_mode=os.path.exists(OUTPUT_FILE))

    print("\n[SUCCESS] STEP 2 COMPLETED — MASTER_DATA READY!")

    if AUTO_MODE:
        print("\n[AUTO] Triggering Data_Analysis_Split.py...")
        try:
            subprocess.run([sys.executable, "Data_Analysis_Split.py", "--auto"], check=True)
            print("[SUCCESS] Step 3 done.")
        except Exception as e:
            print(f"[ERROR] Step 3 failed: {e}")

if __name__ == "__main__":
    main()