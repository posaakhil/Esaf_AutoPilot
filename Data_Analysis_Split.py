import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import sys
import os
import json
import subprocess

# ===== LOAD CONFIG FROM JSON =====
with open("esaf_config.json", 'r', encoding='utf-8') as f:
    config = json.load(f)

ASSIGNEES = config["assignees"]
INDIA_OVERFLOW_THRESHOLD = config["rules"]["india_overflow_threshold"]
MAX_PER_PERSON_DOMESTIC = config["rules"]["max_per_person_domestic"]
MAX_PER_PERSON_MEDITECH = config["rules"]["max_per_person_meditech"]
INDIA_KEYWORD = config["rules"]["india_keyword"]
MEDITECH_KEYWORD = config["rules"]["meditech_keyword"]
REQUESTS_PER_APP_DOMESTIC = config["rules"]["requests_per_app_domestic"]

# ===== CONFIG =====
AUTO_MODE = "--auto" in sys.argv

def load_master_data():
    if not os.path.exists("Today_Assignment.xlsx"):
        print("[ERROR] Today_Assignment.xlsx not found. Run merge_and_cleanup.py first.")
        return None
    try:
        df = pd.read_excel("Today_Assignment.xlsx", sheet_name="Master_Data")
        print(f"[INFO] Loaded {len(df)} rows from Master_Data")
        return df
    except Exception as e:
        print(f"[ERROR] Failed to load Master_Data: {e}")
        return None

def split_data(df):
    india_mask = df['Status'].str.contains(INDIA_KEYWORD, na=False)
    india_df = df[india_mask].copy()
    non_india_df = df[~india_mask].copy()

    meditech_mask = non_india_df['Status'].str.contains(MEDITECH_KEYWORD, na=False)
    meditech_df = non_india_df[meditech_mask].copy()
    domestic_df = non_india_df[~meditech_mask].copy()

    print(f"[INFO] Split: India={len(india_df)}, Domestic={len(domestic_df)}, Meditech={len(meditech_df)}")
    return india_df, domestic_df, meditech_df

def assign_requests_dynamic(india_df, domestic_df, meditech_df, assignees, max_total=15, max_meditech=5, req_per_app=2):
    if india_df.empty and domestic_df.empty and meditech_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    assigned_india = []
    assigned_domestic = []
    assigned_meditech = []
    extra_rows = []
    assignee_counts = {assignee: 0 for assignee in assignees}

    # PHASE 1: INDIA
    if not india_df.empty:
        india_list = india_df.to_dict('records')
        assignee_idx = 0
        for row in india_list:
            if assignee_counts[assignees[assignee_idx]] < max_total:
                row['Assignee'] = assignees[assignee_idx]
                assignee_counts[assignees[assignee_idx]] += 1
                assigned_india.append(row)
            else:
                extra_rows.append(row)
            assignee_idx = (assignee_idx + 1) % len(assignees)

    # PHASE 2: DOMESTIC (APP DIVERSITY)
    if not domestic_df.empty:
        apps = domestic_df['Application'].unique().tolist()
        print(f"[INFO] Discovered {len(apps)} applications: {apps}")

        app_groups = {}
        for app in apps:
            app_groups[app] = domestic_df[domestic_df['Application'] == app].to_dict('records')

        for app in apps:
            group = app_groups[app]
            if not group:
                continue

            print(f"[INFO] Assigning {req_per_app} requests per assignee from '{app}'...")
            assignee_idx = 0

            # Round 1: Assign req_per_app to each assignee
            for _ in range(req_per_app):
                if not group:
                    break
                for assignee in assignees:
                    if assignee_counts[assignee] < max_total and group:
                        row = group.pop(0)
                        row['Assignee'] = assignee
                        assignee_counts[assignee] += 1
                        assigned_domestic.append(row)
                    if not group:
                        break

            # Round 2: Assign leftovers
            while group:
                assignee = assignees[assignee_idx % len(assignees)]
                if assignee_counts[assignee] < max_total:
                    row = group.pop(0)
                    row['Assignee'] = assignee
                    assignee_counts[assignee] += 1
                    assigned_domestic.append(row)
                else:
                    assignee_idx += 1
                    if assignee_idx >= len(assignees) * 2:
                        break
            extra_rows.extend(group)

    # PHASE 3: MEDITECH
    if not meditech_df.empty:
        meditech_list = meditech_df.to_dict('records')
        assignee_idx = 0
        meditech_counts = {assignee: 0 for assignee in assignees}

        for row in meditech_list:
            assigned = False
            start_idx = assignee_idx
            while assignee_idx < start_idx + len(assignees):
                assignee = assignees[assignee_idx % len(assignees)]
                if assignee_counts[assignee] < max_total and meditech_counts[assignee] < max_meditech:
                    row['Assignee'] = assignee
                    assignee_counts[assignee] += 1
                    meditech_counts[assignee] += 1
                    assigned_meditech.append(row)
                    assigned = True
                    break
                assignee_idx += 1

            if not assigned:
                extra_rows.append(row)
            else:
                assignee_idx = (assignee_idx + 1) % len(assignees)

    india_final = pd.DataFrame(assigned_india) if assigned_india else pd.DataFrame()
    domestic_final = pd.DataFrame(assigned_domestic) if assigned_domestic else pd.DataFrame()
    meditech_final = pd.DataFrame(assigned_meditech) if assigned_meditech else pd.DataFrame()
    extra_final = pd.DataFrame(extra_rows) if extra_rows else pd.DataFrame()

    return india_final, domestic_final, meditech_final, extra_final

def apply_sorting_and_styling(worksheet, sort_by_col="Last updated time"):
    sort_col_idx = None
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=col).value == sort_by_col:
            sort_col_idx = col
            break

    if sort_col_idx:
        data_rows = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
            data_rows.append(row)
        try:
            data_rows.sort(key=lambda x: x[sort_col_idx - 1] if x[sort_col_idx - 1] else "")
        except Exception as e:
            print(f"[WARNING] Could not sort by '{sort_by_col}': {e}")

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.value = None

        for r_idx, row_data in enumerate(data_rows, start=2):
            for c_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Styling
    header_fill = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Autofit
    for column in worksheet.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        width = min(max_length + 2, 50)
        worksheet.column_dimensions[letter].width = width

def save_to_sheets(india_df, domestic_df, meditech_df, extra_df):
    domestic_combined = pd.concat([df for df in [domestic_df, meditech_df] if not df.empty], ignore_index=True) if (not domestic_df.empty or not meditech_df.empty) else pd.DataFrame()

    with pd.ExcelWriter("Today_Assignment.xlsx", engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        if not india_df.empty:
            india_df.to_excel(writer, sheet_name="HCA_India", index=False)
        if not domestic_combined.empty:
            domestic_combined.to_excel(writer, sheet_name="HCA_Domestic", index=False)
        if not extra_df.empty:
            extra_df.to_excel(writer, sheet_name="HCA_EXTRA_DATA", index=False)

    wb = load_workbook("Today_Assignment.xlsx")
    for sheet in ["HCA_India", "HCA_Domestic", "HCA_EXTRA_DATA"]:
        if sheet in wb.sheetnames:
            apply_sorting_and_styling(wb[sheet])
    wb.save("Today_Assignment.xlsx")
    print("[SUCCESS] Saved and styled all sheets.")

def main():
    print("[INFO] STEP 3: ULTIMATE DYNAMIC ASSIGNMENT ENGINE")
    print("=" * 50)

    df = load_master_data()
    if df is None:
        return

    india_df, domestic_df, meditech_df = split_data(df)

    if len(india_df) >= INDIA_OVERFLOW_THRESHOLD:
        print(f"[WARNING] India has {len(india_df)} >= {INDIA_OVERFLOW_THRESHOLD} -> moving ALL to HCA_EXTRA_DATA")
        extra_df = india_df
        india_df = pd.DataFrame()
    else:
        extra_df = pd.DataFrame()

    india_assigned, domestic_assigned, meditech_assigned, extra_assigned = assign_requests_dynamic(
        india_df, domestic_df, meditech_df, ASSIGNEES,
        max_total=MAX_PER_PERSON_DOMESTIC,
        max_meditech=MAX_PER_PERSON_MEDITECH,
        req_per_app=REQUESTS_PER_APP_DOMESTIC
    )

    if not extra_assigned.empty:
        extra_df = pd.concat([extra_df, extra_assigned], ignore_index=True) if not extra_df.empty else extra_assigned

    save_to_sheets(india_assigned, domestic_assigned, meditech_assigned, extra_df)

    print("\n[INFO] FINAL ASSIGNMENT SUMMARY:")
    for assignee in ASSIGNEES:
        india_count = len(india_assigned[india_assigned['Assignee'] == assignee]) if not india_assigned.empty else 0
        domestic_count = len(domestic_assigned[domestic_assigned['Assignee'] == assignee]) if not domestic_assigned.empty else 0
        meditech_count = len(meditech_assigned[meditech_assigned['Assignee'] == assignee]) if not meditech_assigned.empty else 0
        total = india_count + domestic_count + meditech_count
        # FIX: Replaced \u2192 with ASCII "->"
        print(f"   {assignee}: India={india_count}, Domestic={domestic_count}, Meditech={meditech_count} -> TOTAL={total}")

    print(f"\n[SUCCESS] STEP 3 COMPLETED — DYNAMIC ASSIGNMENT DONE!")
    print(f"[INFO] HCA_India: {len(india_assigned)} rows")
    print(f"[INFO] HCA_Domestic: {len(domestic_assigned) + len(meditech_assigned)} rows")
    print(f"[INFO] HCA_EXTRA_DATA: {len(extra_df)} rows")

    if AUTO_MODE:
        print("\n[AUTO] Auto-triggering summary_pivot.py...")
        try:
            subprocess.run([sys.executable, "summary_pivot.py", "--auto"], check=True)
            print("[SUCCESS] Step 4 completed automatically.")
        except Exception as e:
            print(f"[ERROR] Failed to auto-trigger summary_pivot.py: {e}")

if __name__ == "__main__":
    main()