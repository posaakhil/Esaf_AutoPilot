# complete_process.py — PyInstaller-safe, no subprocess, error-only pause, clean success flow
import sys
import os
import json
import tempfile
import importlib.util
from pathlib import Path

# ===== PATH & CONFIG HELPERS =====
def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

CONFIG_FILE = "esaf_config.json"
if not os.path.exists(CONFIG_FILE):
    print("[FATAL] ❌ esaf_config.json not found! Place it next to the executable.")
    input("\n[ERROR] Press Enter to exit...")
    sys.exit(1)

with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    config = json.load(f)

# ===== SCRIPT EXECUTION ENGINE =====
def run_script_in_memory(script_name, step_name):
    """Execute script in current process with full error visibility"""
    print(f"\n🚀 [{step_name}] Starting...")
    script_path = get_resource_path(script_name)
    if not os.path.exists(script_path):
        print(f"[FATAL] ❌ Script missing in bundle: {script_name}")
        return False

    try:
        with open(script_path, 'r', encoding='utf-8') as f:
            code = f.read()

        # Prepare globals
        script_globals = {
            '__name__': '__main__',
            '__file__': script_path,
            'sys': sys,
            'os': os,
            'json': json,
            'tempfile': tempfile,
            'Path': Path,
        }

        # Inject common modules
        try:
            import pandas as pd
            script_globals['pd'] = pd
        except ImportError:
            pass
        try:
            from openpyxl import load_workbook
            script_globals['load_workbook'] = load_workbook
        except ImportError:
            pass
        try:
            import plotly.express as px
            import plotly.graph_objects as go
            import plotly.io as pio
            script_globals.update({'px': px, 'go': go, 'pio': pio})
        except ImportError:
            pass

        # Execute
        exec(code, script_globals)
        print(f"✅ [{step_name}] Completed successfully.")
        return True

    except SystemExit as e:
        if e.code == 0:
            print(f"✅ [{step_name}] Exited cleanly.")
            return True
        else:
            print(f"❌ [{step_name}] Exited with error code: {e.code}")
            return False
    except Exception as e:
        print(f"💥 [{step_name}] CRASHED:")
        import traceback
        traceback.print_exc()
        return False

# ===== MAIN WORKFLOW =====
def main():
    print("\n" + "="*80)
    print("🎯 ESAF FULL END-TO-END AUTOMATION (v2.2 - PyInstaller Safe)")
    print("💡 Runs all 5 steps in-memory — no subprocess, works inside .exe")
    print("="*80)

    steps = [
        ("esaf_automation.py", "Step 1: ESAF UI Automation"),
        ("merge_and_cleanup.py", "Step 2: Merge & Cleanup"),
        ("Data_Analysis_Split.py", "Step 3: Assign Requests to Team"),
        ("summary_pivot.py", "Step 4: Summary & Pivot Tables"),
        ("Interactive_Dashboard.py", "Step 5: Interactive Executive Dashboard"),
    ]

    for script, name in steps:
        success = run_script_in_memory(script, name)
        if not success:
            print(f"\n🛑 WORKFLOW FAILED AT: {name}")
            input("\n[DEBUG] Press Enter to view error above and exit...")
            sys.exit(1)

    print("\n" + "="*80)
    print("🎉 ENTIRE ESAF WORKFLOW COMPLETED SUCCESSFULLY!")
    print("📁 Outputs:")
    print("   → Today_Assignment.xlsx (Master_Data, HCA_India, HCA_Domestic, Summary, Pivot)")
    print("   → esaf_executive_dashboard.html")
    print("="*80)
    input("\n[SUCCESS] Press Enter to return to main menu...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user.")
        input("\n[ABORT] Press Enter to return to menu...")
    except Exception as e:
        print(f"\n[FATAL] Unexpected crash: {e}")
        import traceback
        traceback.print_exc()
        input("\n[CRASH] Press Enter to return to menu...")
